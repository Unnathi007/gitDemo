import unittest
import openpyxl
import csvToXlsx
import csv
class test_csvToXlsx(unittest.TestCase):
    def test_checkFileCreation(self):
        filePath = "C:\\Users\\sbatraju\\WebDev\\python\\Test\\thirdQuestion\\csv files\\realestate.csv"
        with open(filePath, "r") as fileReader:
            rows = csv.reader(fileReader)
            numberOfRows = len(list(rows))
        csvToXlsx.csvToxlsx()
        xlsxFilePath = r"C:\Users\sbatraju\WebDev\python\Test\thirdQuestion\csv files\realestate.xlsx"
        xlsxReader = openpyxl.load_workbook(xlsxFilePath)
        sheet = xlsxReader['Sheet1']
        print(numberOfRows,sheet.max_row)
        assert int(sheet.max_row) == numberOfRows
        