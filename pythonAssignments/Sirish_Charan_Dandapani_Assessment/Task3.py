from openpyxl import Workbook
import csv, os, json
import datetime

def create_xlsx_file():
     wb = Workbook()
     ws = wb.active
     with open('realestate.csv', 'r') as f:
          for row in csv.reader(f):
               ws.append(row)
     # print(ws.max_row, ws.max_column)
     # print(os.path.getsize(datetime.datetime.now().strftime("%d_%b_%Y")+'.xlsx'))
     # print(os.path.isfile(datetime.datetime.now().strftime("%d_%b_%Y")+'.xlsx'))
     wb.save(datetime.datetime.now().strftime("%d_%b_%Y")+'.xlsx')
     wb.close()

def create_json_file():
     data = {}
     with open("realestate.csv", encoding='utf-8') as csvf:
          csvReader = csv.DictReader(csvf)
          key = 3787
          for rows in csvReader:
               data[key] = rows
               key += 1
     with open("realestate.json", 'w', encoding='utf-8') as jsonf:
          jsonf.write(json.dumps(data, indent=4))