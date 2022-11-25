import Task3
import unittest
from openpyxl import load_workbook
import datetime
import os

class test_task3_v01(unittest.TestCase):
     def setUp(self):
          Task3.create_xlsx_file()
          self.wb = load_workbook(filename=f'{datetime.datetime.now().strftime("%d_%b_%Y")}'+'.xlsx')
          with open("realestate.csv") as f:
               self.col = len(next(f).split(','))
               self.row = sum(1 for row in f) + 1

     def test_compare_number_of_rows(self):
          ws = self.wb.active
          self.assertEqual( ws.max_row, self.row)
     def test_compare_number_of_columns(self):
          ws = self.wb.active
          self.assertEqual( ws.max_column, self.col)


class test_task3_v02(unittest.TestCase):
     def test_if_file_size(self):
          self.assertTrue(os.path.getsize(datetime.datetime.now().strftime("%d_%b_%Y")+'.xlsx') > 0)

     def test_if_file_exists(self):
          self.assertTrue(os.path.isfile(datetime.datetime.now().strftime("%d_%b_%Y")+'.xlsx'))